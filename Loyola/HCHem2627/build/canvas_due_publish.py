#!/usr/bin/env python3
"""
canvas_due_publish.py — set FALL due dates (per section) and PUBLISH fall
assignments in Honors Chem (course 6624).  SIS sync is intentionally NOT set.

Rules:
  daily homework      -> due 11:59 PM at each section's NEXT meeting
  chapter homework    -> due the next free-response (2nd) exam day
  labs                -> due the next free-response (2nd) exam day
  post-lab quizzes     -> due the section's next meeting AFTER the lab
                          (skip a meeting if it's an exam day)
  chapter quizzes     -> same day as the matching post-lab quiz (next meeting
                          after the lab), else next meeting after the quiz
  practice / reviews  -> no due date
Publishing: every matched fall assignment is published. post_to_sis is left OFF.

USAGE
  Offline preview (no token, reads the inventory file):
      python3 canvas_due_publish.py
  Live preview (read-only, confirms section ids):
      export CANVAS_TOKEN=...; python3 canvas_due_publish.py --live
  Apply (writes due dates + publishes; NO SIS):
      export CANVAS_TOKEN=...; python3 canvas_due_publish.py --live --apply
"""
import os, re, sys, json, datetime, urllib.request, urllib.parse
from openpyxl import load_workbook
try:
    from zoneinfo import ZoneInfo; TZ = ZoneInfo("America/Los_Angeles")
except Exception:
    TZ = None

BASE, COURSE_ID = "https://loyolahs.instructure.com", "6624"
INV = "canvas_ids_6624_renamed.txt"
SRC = "Honors_Chem_2627_Schedule_2.xlsx"
FALL_CUTOFF = datetime.date(2026, 12, 20)
BLANK_ROW_DATE = {50: datetime.date(2026, 12, 15)}
LIVE  = "--live" in sys.argv
APPLY = "--apply" in sys.argv
# map class period -> Canvas section id. Auto-detected in --live; override here if needed:
SECTION_OVERRIDE = {}   # e.g. {5: 123, 6: 124, 7: 125}

# ---------- inventory + matchers (same logic as build_page.py) --------------
A, sec = [], None
for ln in open(INV, encoding="utf-8"):
    if "ASSIGNMENTS" in ln: sec = "A"; continue
    if "FILES" in ln: sec = None; continue
    m = re.match(r"\s*(\d+)\s+(.*?)\s{2,}https?://", ln)
    if m and sec == "A": A.append((int(m.group(1)), m.group(2).strip()))
ID2T = {i: t for i, t in A}
def norm(s): return re.sub(r"\s+", " ", re.sub(r"[^\w\s.#/-]", " ", s.lower())).strip()
TITLES = [(i, t, norm(t)) for i, t in A]
def find(p):
    for i, t, n in TITLES:
        if p(n, t): return i
    return None
def lead(s):
    m = re.match(r"\s*(\d+\.\d+)", s); return m.group(1) if m else None
def m_daily(raw):
    n = norm(raw)
    a = find(lambda x, t: x.rstrip(". ") == n.rstrip(". "))
    if a: return a
    b = re.sub(r"classwork|daily homework", "", n).strip(" ,")
    a = find(lambda x, t: re.sub(r"classwork|daily homework", "", x).strip(" ,") == b and ("classwork" in x or "daily homework" in x))
    if a: return a
    s = lead(raw)
    return find(lambda x, t: s and x.startswith(s) and ("classwork" in x or "daily homework" in x) and "practice" not in x)
def m_homework(text):
    t = text.lower()
    mm = re.search(r"(\d+)\s*-\s*(\d+)", t)
    if mm:
        a, b = mm.groups(); return find(lambda x, tt: "homework" in x and re.search(rf"\b{a}\s*-\s*{b}\b", x))
    m = re.search(r"(?:homework|hw)\s*#?\s*(\d+)(?:\s*part\s*(\d+))?", t)
    if not m: return None
    ch, part = m.groups()
    if part: return find(lambda x, tt: "homework" in x and re.search(rf"(ch\.?|chapter)\s*{ch}\b", x) and f"part {part}" in x)
    return find(lambda x, tt: "homework" in x and re.search(rf"(ch\.?|chapter)\s*{ch}\b", x) and "part" not in x)
def m_quiz(text, postlab):
    m = re.search(r"#\s*(\d+)", text)
    if not m: return find(lambda x, t: "snap" in x and "quiz" in x) if "snap" in text.lower() else None
    n = m.group(1)
    if postlab: return find(lambda x, t: "post" in x and "quiz" in x and re.search(rf"#\s*{n}\b", x))
    return find(lambda x, t: "quiz" in x and "post" not in x and re.search(rf"#\s*{n}\b", x))
def m_lab(text):
    m = re.search(r"(\d+)", text)
    return find(lambda x, t: re.fullmatch(rf"lab {m.group(1)}", x) is not None) if m else None

# ---------- walk schedule, build per-period timelines -----------------------
def pdate(v):
    if v is None: return None
    if isinstance(v, (datetime.datetime, datetime.date)): return v.date() if isinstance(v, datetime.datetime) else v
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", str(v).strip()); return datetime.date(*map(int, m.groups())) if m else None
ws = load_workbook(SRC).active
ROWS = []
for r in range(2, ws.max_row + 1):
    p = {5: pdate(ws.cell(r,2).value), 6: pdate(ws.cell(r,3).value), 7: pdate(ws.cell(r,4).value)}
    if all(v is None for v in p.values()):
        if r in BLANK_ROW_DATE: p = {k: BLANK_ROW_DATE[r] for k in p}
        else: continue
    if any(v is None for v in p.values()) or min(p.values()) > FALL_CUTOFF: continue
    ROWS.append((r, p, ws.cell(r,1).value or ""))
MEET = {P: sorted(rp[P] for _, rp, _ in ROWS) for P in (5,6,7)}
EXAM_DATES = {P: set() for P in (5,6,7)}
FR_EXAM = {P: [] for P in (5,6,7)}      # free-response (2nd) exam days
for _, p, cell in ROWS:
    low = cell.lower()
    if "exam" in low and "review" not in low:
        for P in (5,6,7): EXAM_DATES[P].add(p[P])
        if "part 2" in low or "free response" in low or "final" in low:
            for P in (5,6,7): FR_EXAM[P].append(p[P])
FALL_FINAL = datetime.date(2026, 12, 15)   # fall final day (not an exam row in the .xlsx)
for P in (5,6,7):
    FR_EXAM[P].append(FALL_FINAL); EXAM_DATES[P].add(FALL_FINAL)
    FR_EXAM[P] = sorted(FR_EXAM[P])
LAB_DATE = {P: {} for P in (5,6,7)}     # lab number -> date
for _, p, cell in ROWS:
    for ln in cell.split("\n"):
        s = ln.strip().lower()
        if s.startswith("lab") and "due" not in s:
            n = re.search(r"(\d+)", s)
            if n:
                for P in (5,6,7): LAB_DATE[P][n.group(1)] = p[P]

def nxt_meet(P, d, skip_exam=False):
    for x in MEET[P]:
        if x > d and not (skip_exam and x in EXAM_DATES[P]): return x
    return None
def nxt_fr(P, d):
    for x in FR_EXAM[P]:
        if x >= d: return x
    return None

# ---------- compute due dates ------------------------------------------------
due = {}          # aid -> {period: date}
publish = set()
def setdue(aid, P, d):
    if aid and d: due.setdefault(aid, {})[P] = d
def addpub(aid):
    if aid: publish.add(aid)

for _, p, cell in ROWS:
    lines = [x.strip() for x in cell.split("\n") if x.strip()]
    lect_done = False
    for ln in lines:
        low = ln.lower()
        if "introduction" in low:                      # first-day intro: no due date
            continue
        if "review" in low or ("exam" in low):        # reviews & exams: publish, no computed due here
            continue
        if "quiz" in low:
            postlab = "post" in low
            aid = m_quiz(ln, postlab); addpub(aid)
            n = re.search(r"#\s*(\d+)", ln)
            for P in (5,6,7):
                if postlab and n and n.group(1) in LAB_DATE[P]:
                    d = nxt_meet(P, LAB_DATE[P][n.group(1)], skip_exam=True)
                else:
                    d = nxt_meet(P, p[P], skip_exam=True)
                setdue(aid, P, d)
        elif low.startswith("lab") and "due" not in low:
            aid = m_lab(ln); addpub(aid)
            for P in (5,6,7): setdue(aid, P, nxt_fr(P, p[P]))
        elif "due" in low:
            if low.startswith("lab"):
                addpub(m_lab(ln))                       # lab due handled by the lab row itself
            else:
                aid = m_homework(ln); addpub(aid)       # chapter homework -> FR exam day
                for P in (5,6,7): setdue(aid, P, nxt_fr(P, p[P]))
        else:
            if lect_done: continue
            lect_done = True
            aid = m_daily(ln); addpub(aid)
            paid = None
            s = lead(ln)
            if s:
                pm = re.search(r"part\s*(\d+)", low)
                paid = find(lambda x, t: x.startswith(s) and "practice" in x and (not pm or f"part {pm.group(1)}" in x))
            addpub(paid)                                # practice: publish, no due
            for P in (5,6,7): setdue(aid, P, nxt_meet(P, p[P]))
# fall exams + reviews: publish them too
for i, t in A:
    tl = t.lower()
    if re.search(r"\bexam\s*[123]\b", tl) or "semester 1 final" in tl or ("exam" in tl and "review" in tl and re.search(r"[123]", tl)):
        publish.add(i)

# ---------- report -----------------------------------------------------------
def fmt(d): return d.strftime("%a %-m/%-d") if d else "—"
print(f"COURSE {COURSE_ID} — DUE DATES (11:59 PM){' — DRY RUN' if not APPLY else ''}\n")
by_kind = sorted(due.items(), key=lambda kv: ID2T[kv[0]])
for aid, dd in by_kind:
    per = "  ".join(f"P{P} {fmt(dd.get(P))}" for P in (5,6,7))
    print(f"  {ID2T[aid][:44]:44} | {per}")
print(f"\nAssignments with due dates: {len(due)}   |   Assignments to publish: {len(publish)}")
print("SIS sync: OFF (post_to_sis will NOT be set)")

import json as _json
_items = sorted(due.items(), key=lambda kv: ID2T.get(kv[0], ""))
_out = {str(a): {"_title": ID2T.get(a, ""), **{str(P): d.isoformat() for P, d in dd.items()}} for a, dd in _items}
with open("due_dates.json", "w") as _f:
    _json.dump(_out, _f, indent=1)
print("wrote due_dates.json")

# ---------- apply ------------------------------------------------------------
def api(path, method="GET", body=None):
    tok = os.environ["CANVAS_TOKEN"]
    data = urllib.parse.urlencode(body, doseq=True).encode() if body else None
    req = urllib.request.Request(f"{BASE}/api/v1/{path}", data=data, method=method,
                                 headers={"Authorization": f"Bearer {tok}"})
    out, url = [], req
    while url:
        with urllib.request.urlopen(url) as r:
            j = json.loads(r.read()); link = r.headers.get("Link", "")
        out += j if isinstance(j, list) else [j]
        m = re.search(r'<([^>]+)>;\s*rel="next"', link)
        url = urllib.request.Request(m.group(1), headers={"Authorization": f"Bearer {tok}"}) if (m and method=="GET") else None
    return out

def due_at(d):
    dt = datetime.datetime(d.year, d.month, d.day, 23, 59, tzinfo=TZ) if TZ else datetime.datetime(d.year, d.month, d.day, 23, 59)
    return dt.isoformat()

if LIVE:
    secs = api(f"courses/{COURSE_ID}/sections?per_page=100")
    pmap = dict(SECTION_OVERRIDE)
    for s in secs:
        for P in (5,6,7):
            if P not in pmap and re.search(rf"\b{P}\b|period\s*{P}", s["name"].lower()): pmap[P] = s["id"]
    print("\nSection map (period -> id):", pmap)
    for s in secs: print(f"   section {s['id']}: {s['name']}")
    missing = [P for P in (5,6,7) if P not in pmap]
    if missing: sys.exit(f"Could not map periods {missing} to sections — set SECTION_OVERRIDE at top of file.")
    if APPLY:
        print("\nAPPLYING ...")
        for aid, dd in due.items():
            for P, d in dd.items():
                api(f"courses/{COURSE_ID}/assignments/{aid}/overrides", "POST",
                    {"assignment_override[course_section_id]": pmap[P],
                     "assignment_override[due_at]": due_at(d)})
            print(f"  due set: {ID2T[aid][:40]}")
        for aid in publish:
            api(f"courses/{COURSE_ID}/assignments/{aid}", "PUT", {"assignment[published]": "true"})
        print(f"published {len(publish)} assignments. SIS left OFF. done.")
    else:
        print("\n(--live preview only; add --apply to write)")

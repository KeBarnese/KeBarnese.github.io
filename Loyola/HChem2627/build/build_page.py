#!/usr/bin/env python3
"""
build_page.py (v2) — writes index.html directly from due_dates.json (all
"due"/"quiz" pills) and lecture_pages.json (all "lect"/"lab"/"exam" occurrence
pills). The schedule .xlsx is READ ONLY to validate that every date/period
combo used below is actually a day that section meets -- it is no longer
parsed for any assignment/lecture text. Column A of the sheet can hold
anything (old free text, or just a school-day number) without affecting the
build at all.

  python3 build_page.py

Inputs:
  * INV   : canvas_ids_6624_renamed.txt      (Canvas title/id inventory, used
            only to find each daily-homework's "practice" sibling assignment)
  * SRC   : the schedule .xlsx               (cols B/C/D = period 5/6/7 dates;
            used only to sanity-check dates below, never to place events)
  * due_dates.json      : source of truth for ALL due/quiz pills
  * lecture_pages.json  : source of truth for ALL lecture/lab/exam pills
  * TPL   : index_template.html              (placeholders __COURSE__/__AIDS__/__EVENTS__)
Output:
  * index.html

See assignment_kind.py for the title -> kind classifier. If Canvas due dates
change, re-run canvas_pull_due.py then this script -- no spreadsheet edit
needed. If a lecture/lab/exam moves, edit lecture_pages.json (or use
lecture_pages_cli.py) then re-run this script -- also no spreadsheet edit.
"""
import re, os, sys, json, datetime, pathlib
from openpyxl import load_workbook
from assignment_kind import classify

# ---- CONFIG ----------------------------------------------------------------
COURSE_ID   = "6624"
INV         = "canvas_ids_6624_renamed.txt"
SRC         = "Honors_Chem_2627_Schedule_2.xlsx"
TPL         = "index_template.html"
OUT         = "../index.html"   # written to the course root (run this from build/)
FALL_CUTOFF = datetime.date(2026, 12, 20)
DUE_JSON    = "due_dates.json"        # source of truth for due/quiz pills
PAGES       = "lecture_pages.json"    # source of truth for lecture/lab/exam pills

# Exam-review PDFs are hand-placed (see EXTRAS in index_template.html) --
# these four titles are matched by EXACT title against the Canvas file/page
# inventory below, same mechanism as before.
REVIEW_TITLES = {"1": "Exam 1 Review - Build On",
                 "2": "Exam 2 Review - Build On",
                 "3": "Exam 3, Chapter 4 review, Build On",
                 "F": "Semester 1 Final Review"}

# ---- parse Canvas title inventory (only used for practice-link matching
#      and for resolving the four review-PDF ids above) --------------------
A, sec = [], None
for ln in open(INV, encoding="utf-8"):
    if "ASSIGNMENTS" in ln: sec = "A"; continue
    if "FILES" in ln: sec = None; continue
    if ln.startswith("===") or "CLASSIC QUIZZES" in ln: continue
    m = re.match(r"\s*(\d+)\s+(.*?)\s{2,}https?://", ln)
    if m and sec == "A": A.append((int(m.group(1)), m.group(2).strip()))
ID2T = {i: t for i, t in A}
def norm(s): return re.sub(r"\s+", " ", re.sub(r"[^\w\s.#/-]", " ", s.lower())).strip()
TITLES = [(i, t, norm(t)) for i, t in A]
def find(pred):
    for i, t, n in TITLES:
        if pred(n, t): return i
    return None
def lead_section(s):
    m = re.match(r"\s*(\d+\.\d+)", s); return m.group(1) if m else None
def _part_of(s):
    m = re.search(r"part\s*(\d+)", s or "", re.I)
    return m.group(1) if m else None

def m_practice(raw):
    """Practice copy for a daily-homework title. Same matching rules as the
    old spreadsheet-text version, just fed the due_dates.json _title instead."""
    s = lead_section(raw)
    if not s:
        return None
    want_part = _part_of(raw)
    if want_part:
        aid = find(lambda x, t: x.startswith(s) and "practice" in x
                   and _part_of(x) == want_part)
        if aid:
            return aid
    aid = find(lambda x, t: x.startswith(s) and "practice" in x
               and _part_of(x) is None)
    if aid:
        return aid
    return None

def m_review(n):
    want = norm(REVIEW_TITLES.get(n, "")); return find(lambda x, t: x == want)

def pdate(v):
    if v is None: return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.date() if isinstance(v, datetime.datetime) else v
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", str(v).strip())
    return datetime.date(int(m[1]), int(m[2]), int(m[3])) if m else None
def iso(d):
    return d.strftime("%Y-%m-%d") if isinstance(d, (datetime.date, datetime.datetime)) else d

AID, events, warnings = {}, [], []
def key_for(aid):
    if aid is None: return None
    k = f"A{aid}"
    AID[k] = aid
    return k

# ---- calendar validity check (xlsx cols B/C/D only -- never column A) ----
# Builds {period: set(iso date)} so every due_dates.json / lecture_pages.json
# date can be sanity-checked against the actual calendar. Non-fatal: a bad
# date becomes a warning, not a build failure.
VALID = {5: set(), 6: set(), 7: set()}
if os.path.exists(SRC):
    ws = load_workbook(SRC).active
    for r in range(2, ws.max_row + 1):
        for col, per in ((2, 5), (3, 6), (4, 7)):
            d = pdate(ws.cell(r, col).value)
            if d is not None:
                VALID[per].add(iso(d))
else:
    warnings.append(("CALENDAR?", f"{SRC} not found -- skipping date sanity checks"))

def check_date(period, d, label):
    if VALID[5] or VALID[6] or VALID[7]:  # only check if we actually loaded a calendar
        if d not in VALID[period]:
            warnings.append(("DATE?", f"{label}: {d} is not a period-{period} school day per {SRC}"))

def groups_from_periods(per_dates):
    """per_dates: {'5': iso, '6': iso, '7': iso} (missing periods allowed).
    Returns list of (date, periods) grouped so equal dates collapse to p=0
    when ALL three periods share it, else a sorted period list."""
    g = {}
    for p_str in ("5", "6", "7"):
        d = per_dates.get(p_str)
        if d:
            g.setdefault(d, []).append(int(p_str))
    all_equal = len(g) == 1 and set(next(iter(g.values()))) == {5, 6, 7}
    out = []
    for d, pers in g.items():
        out.append((d, 0 if all_equal else sorted(pers)))
    return out

# ---- 1. due/quiz pills, entirely from due_dates.json -----------------------
DUE = json.load(open(DUE_JSON)) if os.path.exists(DUE_JSON) else {}

def chapter_hw_label(title):
    m = re.search(r"chapter\s*(\d+(?:\s*-\s*\d+)?)\s*homework(?:\s*,?\s*part\s*(\d+))?", title, re.I)
    if not m:
        return title.strip() + " due"
    ch, part = m.group(1), m.group(2)
    lbl = f"Chapter HW {ch}"
    if part:
        lbl += f" part {part}"
    return lbl + " due"

def daily_hw_label(title):
    stripped = re.sub(r"\s*,?\s*daily homework\s*$", "", title, flags=re.I).strip(" ,")
    return "DHW: " + stripped

for aid, entry in DUE.items():
    title = entry.get("_title", "")
    kind = entry.get("kind") or classify(title)
    per_dates = {p: entry[p] for p in ("5", "6", "7") if p in entry}
    if not per_dates:
        continue

    hw2 = None
    if kind == "daily_homework":
        label = daily_hw_label(title)
        paid = m_practice(title)
        hw2 = key_for(paid) if paid else None
        k = "due"
    elif kind == "chapter_homework":
        label = chapter_hw_label(title)
        k = "due"
    elif kind == "lab_due":
        label = title.strip() + " due"
        k = "due"
    elif kind in ("quiz", "postlab_quiz"):
        label = title
        k = "quiz"
    elif kind == "info":
        label = title
        k = "info"
    else:  # "other" -- unclassified, still shown, still flagged
        label = title.strip() + " due"
        k = "due"
        warnings.append(("KIND?", f"{aid} ({title!r}) fell through to 'other' -- "
                                    f"check assignment_kind.classify() or add an override"))

    hw = key_for(aid)
    for d, pers in groups_from_periods(per_dates):
        for p in ([5, 6, 7] if pers == 0 else pers):
            check_date(p, d, f"due_dates.json {aid} ({title})")
        ev = {"d": d, "p": pers, "k": k, "t": label, "hw": hw}
        if hw2:
            ev["hw2"] = hw2
        events.append(ev)

# ---- 2. lecture/lab/exam occurrence pills, entirely from lecture_pages.json
PAGE_URLS = {}
if os.path.exists(PAGES):
    PG = json.load(open(PAGES))
else:
    PG = {}
    warnings.append(("PAGES?", f"{PAGES} not found -- no lecture/lab/exam pills will be built"))

KIND2K = {"lecture": "lect", "lab": "lab", "exam": "exam"}
for pid, entry in PG.items():
    if not isinstance(entry, dict):
        continue
    etype = entry.get("type")
    k = KIND2K.get(etype)
    if k is None:
        warnings.append(("PAGETYPE?", f"{pid} has unknown type {etype!r} -- skipped"))
        continue
    dates = entry.get("dates") or {}
    per_dates = {p: dates[p] for p in ("5", "6", "7") if dates.get(p)}
    if not per_dates:
        warnings.append(("NODATE", f"{pid} ({entry.get('title')}) has no dates set -- skipped"))
        continue
    if entry.get("page_url"):
        PAGE_URLS[pid] = entry["page_url"]
    for d, pers in groups_from_periods(per_dates):
        for p in ([5, 6, 7] if pers == 0 else pers):
            check_date(p, d, f"{PAGES} {pid} ({entry.get('title')})")
        ev = {"d": d, "p": pers, "k": k, "t": entry.get("title", pid), "hw": None}
        if pid in PAGE_URLS:
            ev["pg"] = pid
        events.append(ev)

events.sort(key=lambda e: (e["d"], {"info":0,"lect":1,"exam":2,"quiz":3,"lab":4,"due":5,"rev":6}.get(e["k"], 9)))

# ---- assemble the page -----------------------------------------------------
events_js = "const EVENTS = " + json.dumps(events, separators=(",", ":")) \
    .replace("},{", "},\n  {").replace("[{", "[\n  {").replace("}]", "}\n];") + ";"
aids_js = "const ASSIGNMENT_IDS = {\n" + "".join(f"  '{k}': {v},\n" for k, v in AID.items()) + "};"
pages_js = "const PAGE_URLS = {\n" + "".join(
    f"  '{k}': '{v}',\n" for k, v in sorted(PAGE_URLS.items())) + "};"

tpl = pathlib.Path(TPL).read_text()
tpl = tpl.replace("__COURSE__", COURSE_ID).replace("__AIDS__", aids_js) \
         .replace("__PAGES__", pages_js).replace("__EVENTS__", events_js)

# point the hand-placed EXTRAS reviews (all four, including Exam 1) at the
# matched Canvas file/page ids
for rn, placeholder in (("1", "'Exam_1_Review_Build_On'"),
                        ("2", "'Exam_2_Review_Build_On'"),
                        ("3", "'Exam_3_Chapter_4_review_Build_On'"),
                        ("F", "'Semester_1_Final_Review'")):
    aid = m_review(rn)
    if aid is None:
        warnings.append(("REVIEW?", f"Exam {rn} review PDF title not found in {INV}"))
        continue
    tpl = tpl.replace(placeholder, f"'A{aid}'")
    key_for(aid)

pathlib.Path(OUT).write_text(tpl)

print(f"course {COURSE_ID}: {len(events)} events, {len(AID)} assignment ids matched, "
      f"{len(PAGE_URLS)} lecture/lab/exam pages linked")
print(f"WARNINGS ({len(warnings)}):")
for tag, t in warnings:
    print(f"  [{tag}] {t}")

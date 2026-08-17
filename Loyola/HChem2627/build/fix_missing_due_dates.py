#!/usr/bin/env python3
"""
fix_missing_due_dates.py — set due dates (periods 5, 6, 7) for the 6 daily
homeworks that canvas_due_publish.py missed, in Honors Chem (course 6624).

Why these six were missed: canvas_due_publish.py only reads the FIRST
classwork/daily-homework line out of each schedule-row cell (see its
`lect_done` flag). These six rows each teach a SECOND section on the same
day (e.g. row 11 covers "2.2-2.5 part 2" AND "2.1 part 1" on the same day),
so the second daily homework in the cell never got matched or given a due
date. This script hardcodes those six assignment ids + their schedule row,
computes "due at the section's next meeting after that row's date" (same
rule as daily homework in canvas_due_publish.py), and (optionally) writes
the overrides to Canvas. SIS sync is intentionally left OFF, same as
canvas_due_publish.py. Nothing is published/unpublished by this script.

USAGE
  Preview only (no token needed):
      python3 fix_missing_due_dates.py
  Live preview (confirms section ids, still read-only):
      export CANVAS_TOKEN=...; python3 fix_missing_due_dates.py --live
  Apply (writes the 18 overrides = 6 assignments x 3 sections):
      export CANVAS_TOKEN=...; python3 fix_missing_due_dates.py --live --apply

After applying, this also rewrites due_dates.json (merging in these six)
so the schedule page stays in sync — run `python3 build_page.py` afterward.
"""
import os, re, sys, json, datetime, urllib.request, urllib.parse
from openpyxl import load_workbook
try:
    from zoneinfo import ZoneInfo; TZ = ZoneInfo("America/Los_Angeles")
except Exception:
    TZ = None

BASE, COURSE_ID = "https://loyolahs.instructure.com", "6624"
SRC = "Honors_Chem_2627_Schedule_2.xlsx"
DUE_JSON = "due_dates.json"
FALL_CUTOFF = datetime.date(2026, 12, 20)
BLANK_ROW_DATE = {50: datetime.date(2026, 12, 15)}
LIVE = "--live" in sys.argv
APPLY = "--apply" in sys.argv
# map class period -> Canvas section id. Auto-detected in --live; override here if needed:
SECTION_OVERRIDE = {}   # e.g. {5: 123, 6: 124, 7: 125}

# assignment id -> (schedule row, title) -- the schedule row is where each
# homework's lecture actually falls, per canvas_ids_6624_renamed.txt / the
# .xlsx (see script docstring for why these rows weren't auto-matched).
TARGETS = {
    266475: (11, "2.1, atomic structure part 1, Daily Homework"),
    266482: (13, "2.6, Nomenclature part 2."),
    266499: (21, "3.5 part 2, Daily Homework"),
    266506: (28, "4.1 and 4.3 Daily Homework"),
    266508: (29, "4.2 part 2, double displacement Daily Homework"),
    266533: (48, "9.3 Daily Homework"),
}

# ---------- read schedule, build per-period meeting calendars ---------------
def pdate(v):
    if v is None: return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.date() if isinstance(v, datetime.datetime) else v
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", str(v).strip())
    return datetime.date(*map(int, m.groups())) if m else None

ws = load_workbook(SRC).active
ROWS = {}   # row number -> {5: date, 6: date, 7: date}
for r in range(2, ws.max_row + 1):
    p = {5: pdate(ws.cell(r, 2).value), 6: pdate(ws.cell(r, 3).value), 7: pdate(ws.cell(r, 4).value)}
    if all(v is None for v in p.values()):
        if r in BLANK_ROW_DATE: p = {k: BLANK_ROW_DATE[r] for k in p}
        else: continue
    if any(v is None for v in p.values()) or min(p.values()) > FALL_CUTOFF: continue
    ROWS[r] = p
MEET = {P: sorted(rp[P] for rp in ROWS.values()) for P in (5, 6, 7)}

def nxt_meet(P, d):
    for x in MEET[P]:
        if x > d: return x
    return None

missing_rows = [row for _, (row, _) in TARGETS.items() if row not in ROWS]
if missing_rows:
    sys.exit(f"Schedule row(s) {missing_rows} not found in {SRC} (or fall outside FALL_CUTOFF) "
              f"-- check the .xlsx hasn't shifted rows.")

due = {}   # aid -> {period: date}
for aid, (row, _title) in TARGETS.items():
    p = ROWS[row]
    due[aid] = {P: nxt_meet(P, p[P]) for P in (5, 6, 7)}

# ---------- report -----------------------------------------------------------
def fmt(d): return d.strftime("%a %-m/%-d") if d else "--"
print(f"COURSE {COURSE_ID} -- MISSING DAILY HOMEWORK DUE DATES (11:59 PM)"
      f"{' -- DRY RUN' if not APPLY else ''}\n")
for aid, (row, title) in TARGETS.items():
    dd = due[aid]
    per = "  ".join(f"P{P} {fmt(dd.get(P))}" for P in (5, 6, 7))
    print(f"  {aid}  {title[:44]:44} | row {row:>3} | {per}")
print(f"\n{len(due)} assignments x 3 sections = {len(due) * 3} overrides to write.")
print("SIS sync: OFF (post_to_sis will NOT be set). Nothing is published/unpublished.")

# ---------- merge into due_dates.json ----------------------------------------
if os.path.exists(DUE_JSON):
    with open(DUE_JSON) as f:
        existing = json.load(f)
else:
    existing = {}
for aid, (_row, title) in TARGETS.items():
    entry = existing.setdefault(str(aid), {"_title": title})
    entry["_title"] = title
    for P, d in due[aid].items():
        if d: entry[str(P)] = d.isoformat()
with open(DUE_JSON, "w") as f:
    json.dump(existing, f, indent=1)
print(f"\nwrote {DUE_JSON} (merged in {len(TARGETS)} entries). "
      f"Run `python3 build_page.py` to refresh the schedule page.")

# ---------- apply to Canvas ---------------------------------------------------
def api(path, method="GET", body=None):
    tok = os.environ["CANVAS_TOKEN"]
    data = urllib.parse.urlencode(body, doseq=True).encode() if body else None
    req = urllib.request.Request(f"{BASE}/api/v1/{path}", data=data, method=method,
                                  headers={"Authorization": f"Bearer {tok}"})
    out, url = [], req
    while url:
        with urllib.request.urlopen(url) as r:
            j = json.loads(r.read()); link = r.headers.get("Link", "")
        out += j if isinstance(j, list) else [j]
        m = re.search(r'<([^>]+)>;\s*rel="next"', link)
        url = urllib.request.Request(m.group(1), headers={"Authorization": f"Bearer {tok}"}) if (m and method == "GET") else None
    return out

def due_at(d):
    dt = datetime.datetime(d.year, d.month, d.day, 23, 59, tzinfo=TZ) if TZ else datetime.datetime(d.year, d.month, d.day, 23, 59)
    return dt.isoformat()

if LIVE:
    if "CANVAS_TOKEN" not in os.environ:
        sys.exit("\n--live requires a Canvas token: export CANVAS_TOKEN=... first "
                  "(the offline report above and due_dates.json were still written).")
    secs = api(f"courses/{COURSE_ID}/sections?per_page=100")
    pmap = dict(SECTION_OVERRIDE)
    for s in secs:
        for P in (5, 6, 7):
            if P not in pmap and re.search(rf"\b{P}\b|period\s*{P}", s["name"].lower()): pmap[P] = s["id"]
    print("\nSection map (period -> id):", pmap)
    for s in secs: print(f"   section {s['id']}: {s['name']}")
    missing = [P for P in (5, 6, 7) if P not in pmap]
    if missing: sys.exit(f"Could not map periods {missing} to sections -- set SECTION_OVERRIDE at top of file.")

    if APPLY:
        print("\nAPPLYING ...")
        for aid, dd in due.items():
            for P, d in dd.items():
                if not d: continue
                api(f"courses/{COURSE_ID}/assignments/{aid}/overrides", "POST",
                    {"assignment_override[course_section_id]": pmap[P],
                     "assignment_override[due_at]": due_at(d)})
            print(f"  due set: {aid}  {TARGETS[aid][1][:40]}")
        print("done. SIS left OFF.")
    else:
        print("\n(--live preview only; add --apply to write)")

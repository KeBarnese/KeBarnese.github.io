#!/usr/bin/env python3
"""
patch_schedule_sections.py - rewrite the section numbers inside
Honors_Chem_2627_Schedule_2.xlsx to match the textbook, chapters 1-9.

    python3 patch_schedule_sections.py                       # preview
    python3 patch_schedule_sections.py --apply               # rewrite in place
    python3 patch_schedule_sections.py --apply --out new.xlsx

Only column A text changes; every date, format, row height and column width is
left exactly as it was.  Run this in the same sitting as
canvas_rename_sections.py - build_page.py matches the two by section number.
"""
import argparse, os, sys, shutil
from openpyxl import load_workbook
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from renames import SCHED

SRC = "Honors_Chem_2627_Schedule_2.xlsx"

def main():
    p = argparse.ArgumentParser()
    p.add_argument("--apply", action="store_true")
    p.add_argument("--src", default=SRC)
    p.add_argument("--out")
    a = p.parse_args()

    wb = load_workbook(a.src); ws = wb.active
    changes = []
    for row in ws.iter_rows(min_col=1, max_col=1):
        c = row[0]
        if not isinstance(c.value, str): continue
        out = []
        for ln in c.value.split("\n"):
            new = ln
            for old, rep in SCHED:
                if ln.strip() == old:
                    new = ln.replace(old, rep)
                    changes.append((c.row, ln, new)); break
            out.append(new)
        c.value = "\n".join(out)

    print(f"\n{len(changes)} schedule lines would change\n" + "-"*78)
    for r, old, new in changes:
        print(f"  row {r:>3}   {old}\n           -> {new}")
    if not a.apply:
        print("\n(dry run - re-run with --apply to write)")
        return
    dest = a.out or a.src
    if dest == a.src:
        shutil.copy(a.src, a.src + ".bak")
        print(f"\nbacked up original to {a.src}.bak")
    wb.save(dest)
    print(f"wrote {dest}")

if __name__ == "__main__":
    main()

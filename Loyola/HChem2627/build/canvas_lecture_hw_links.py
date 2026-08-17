#!/usr/bin/env python3
"""
canvas_lecture_hw_links.py — put the daily homework link (and its practice
version, when one exists) on each lecture PAGE in Honors Chemistry (course 6624).

This runs in TWO PHASES on purpose.  Your lecture titles and your Canvas
assignment names use different section numbering in several places (lecture
"1.7 Dimensional Analysis" vs assignment "1.6, Dimensional analysis"), and a
few lectures/assignments genuinely don't pair 1-to-1.  So the script proposes a
mapping for you to eyeball, and only writes to Canvas once you're happy with it.

    PHASE 1 — propose
        export CANVAS_TOKEN='...'
        python3 canvas_lecture_hw_links.py --propose

      Writes lecture_hw_map.json and prints a review table.  Open that file and
      fix any row flagged "review" or "none": set "assignment_id" (and
      "practice_id") to the right Canvas id, or set "skip": true to leave that
      page alone.

      Plain --propose OVERWRITES the file (a .bak copy is kept).  Once you have
      edited the map, use --propose --merge instead: it keeps your
      assignment_id / skip / confidence choices and only refreshes the practice
      links, so adding practice assignments in Canvas later does not cost you
      your review work.

    PHASE 2 — apply
        python3 canvas_lecture_hw_links.py                 # dry run: show the HTML
        python3 canvas_lecture_hw_links.py --apply         # write to Canvas
        python3 canvas_lecture_hw_links.py --apply --verify

IDEMPOTENT: the block is wrapped in <!-- daily-hw:start --> / <!-- daily-hw:end -->
markers.  Re-running replaces the existing block instead of stacking duplicates,
so it is safe to run again after you fix a mapping.  Nothing else on the page is
touched, and page publish state is never changed.

Run from inside the build/ folder — it reads notes/map.json, lecture_pages.json,
and Honors_Chem_2627_Schedule_2.xlsx the same way build_page.py does.
"""

import os
import re
import sys
import json
import time
import html
import difflib
import argparse
import urllib.request
import urllib.parse
import urllib.error

BASE = "https://loyolahs.instructure.com"
COURSE_ID = "6624"
PER_PAGE = 100

MAP_JSON = "notes/map.json"
PAGES_JSON = "lecture_pages.json"
SCHEDULE_XLSX = "Honors_Chem_2627_Schedule_2.xlsx"
OUT_MAP = "lecture_hw_map.json"

START = "<!-- daily-hw:start -->"
END = "<!-- daily-hw:end -->"

# Assignments that ARE daily homework but whose titles don't say so.
# (The three 2.6 Nomenclature ones — see the note in the review output.)
EXTRA_DAILY_IDS = {266481, 266482, 266483}

# schedule lines that are never a lecture's daily homework
SKIP_LINE = re.compile(
    r'^(quiz|post.?lab|lab\b|exam|.*\bdue\b|.*review|introduction|snap)', re.I)


# ---- API --------------------------------------------------------------------
def api(path, method="GET", body=None, params=None):
    token = os.environ.get("CANVAS_TOKEN")
    if not token:
        sys.exit("ERROR: set CANVAS_TOKEN first  ->  export CANVAS_TOKEN='...'")
    url = f"{BASE}/api/v1/{path.lstrip('/')}"
    if params:
        url += "?" + urllib.parse.urlencode(params, doseq=True)
    data = urllib.parse.urlencode(body, doseq=True).encode() if body else None
    out = []
    while url:
        req = urllib.request.Request(
            url, data=data, method=method,
            headers={"Authorization": f"Bearer {token}"})
        with urllib.request.urlopen(req) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
            link = resp.headers.get("Link", "")
        out.extend(payload if isinstance(payload, list) else [payload])
        nxt = None
        if method == "GET":
            for part_ in link.split(","):
                m = re.search(r'<([^>]+)>;\s*rel="next"', part_)
                if m:
                    nxt = m.group(1)
                    break
        url = nxt
        data = None
    return out


# ---- text helpers used for matching ----------------------------------------
def denum(s):
    """02.1 -> 2.1 so leading-zero titles compare correctly."""
    return re.sub(r'\b0+(\d)', r'\1', s)


def norm(s):
    s = denum((s or "").lower())
    s = re.sub(r'classwork|daily homework|practice|naming inorganic compounds', '', s)
    s = re.sub(r'[^\w\s.#-]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip(' ,.-')


def sections(s):
    return set(re.findall(r'\d+\.\d+', denum(s or "")))


def part_no(s):
    m = re.search(r'part\s*(\d+)', s or "", re.I)
    return m.group(1) if m else None


def score(a, b):
    """Similarity between a schedule/lecture string and an assignment name.
    Returns (score, sections_matched_exactly)."""
    s = difflib.SequenceMatcher(None, norm(a), norm(b)).ratio()
    sa, sb = sections(a), sections(b)
    exact = bool(sa) and sa == sb
    if sa and sb:
        s += 0.40 if exact else (0.20 if sa & sb else -0.45)
    pa, pb = part_no(a), part_no(b)
    if pa and pb:
        s += 0.30 if pa == pb else -0.55
    return s, exact


# ---- phase 1: propose -------------------------------------------------------
def propose(args):
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("ERROR: --propose needs openpyxl (canvas_due_publish.py uses it too).\n"
                 "       pip3 install openpyxl")

    lectures = json.load(open(MAP_JSON, encoding="utf-8"))
    pages = json.load(open(PAGES_JSON, encoding="utf-8"))
    ws = load_workbook(SCHEDULE_XLSX, data_only=True).active

    assignments = api(f"courses/{args.course}/assignments", params={"per_page": PER_PAGE})
    daily, practice = [], []
    for a in assignments:
        name = a.get("name", "")
        low = name.lower()
        if "practice" in low:
            practice.append((a["id"], name))
        elif re.search(r"daily\s+homework", low) or a["id"] in EXTRA_DAILY_IDS:
            daily.append((a["id"], name))
    D, P = dict(daily), dict(practice)
    ALL = {a["id"]: a.get("name", "") for a in assignments}
    print(f"# pulled {len(assignments)} assignments: {len(daily)} daily homework, "
          f"{len(practice)} practice\n", file=sys.stderr)

    # schedule row -> candidate daily assignments
    row_hits = {}
    for r in range(2, ws.max_row + 1):
        hits = []
        for line in str(ws.cell(r, 1).value or "").split("\n"):
            line = line.strip()
            if not line or SKIP_LINE.match(line):
                continue
            best = (0.0, False, None)
            for aid, an in daily:
                s, exact = score(line, an)
                if s > best[0]:
                    best = (s, exact, aid)
            s, exact, aid = best
            if aid and s >= (0.70 if exact else 0.85):
                hits.append((aid, round(s, 2)))
        if hits:
            row_hits[r] = hits

    # practice partner for a daily assignment
    def practice_for(an):
        best = (0.0, None)
        for pid, pn in practice:
            s = difflib.SequenceMatcher(None, norm(an), norm(pn)).ratio()
            sa, sp = sections(an), sections(pn)
            if sa and sa == sp:
                s += 0.30
            elif sa and sp and not (sa & sp):
                s -= 0.50
            pa, pp = part_no(an), part_no(pn)
            if pa and pp and pa != pp:
                s -= 0.80
            if s > best[0]:
                best = (s, pid)
        return best[1] if best[0] >= 0.80 else None

    out, claims = {}, {}
    for L in lectures:
        lid, row, title = L["id"], L["row"], L["title"]
        page = pages.get(lid, {})
        cands = row_hits.get(row, [])
        entry = {
            "lecture_title": title,
            "page_url": page.get("page_url"),
            "schedule_row": row,
            "assignment_id": None, "assignment_name": None,
            "practice_id": None, "practice_name": None,
            "confidence": "none",
            "notes": "",
            "alternatives": [],
            "skip": False,
        }
        if not page.get("page_url"):
            entry["notes"] = "no Canvas page in lecture_pages.json — cannot write"
            out[lid] = entry
            continue
        if not cands:
            entry["notes"] = (f"no daily-homework line found in schedule row {row} — "
                              f"this lecture may genuinely have no homework")
            out[lid] = entry
            continue

        ranked = []
        for aid, rs in cands:
            s, _ = score(title, D[aid])
            if part_no(title) and part_no(title) == part_no(D[aid]):
                s += 0.5
            ranked.append((s, aid, rs))
        ranked.sort(reverse=True)
        aid = ranked[0][1]
        entry["assignment_id"] = aid
        entry["assignment_name"] = D[aid]
        entry["confidence"] = "review" if len(cands) > 1 else "high"
        if len(cands) > 1:
            entry["alternatives"] = [[a, D[a]] for _, a, _ in ranked[1:]]
            entry["notes"] = (f"{len(cands)} homework assignments sit in schedule row {row} — "
                              f"confirm this is the right one")
        pid = practice_for(D[aid])
        if pid:
            entry["practice_id"], entry["practice_name"] = pid, P[pid]
        else:
            entry["notes"] = (entry["notes"] + "; " if entry["notes"] else "") + \
                             "no practice version found for this homework"
        out[lid] = entry
        claims.setdefault(aid, []).append(lid)

    # cross-checks
    for aid, lids in claims.items():
        if len(lids) > 1:
            for lid in lids:
                out[lid]["confidence"] = "review"
                out[lid]["notes"] = (out[lid]["notes"] + "; " if out[lid]["notes"] else "") + \
                    f"this assignment is also linked from {', '.join(x for x in lids if x != lid)}"

    # ---- merge: keep your hand edits, refresh only the practice links --------
    merge_report = []
    if args.merge and os.path.exists(OUT_MAP):
        prev = json.load(open(OUT_MAP, encoding="utf-8"))
        merged = {}
        for lid, fresh in out.items():
            old = prev.get(lid)
            if not old:
                fresh["notes"] = ("new lecture not in the previous map; "
                                  + fresh["notes"]).strip("; ")
                merged[lid] = fresh
                merge_report.append((lid, "NEW", "added to the map"))
                continue

            # everything you may have curated by hand is preserved verbatim
            e = dict(old)
            e["page_url"] = fresh["page_url"] or old.get("page_url")
            e["lecture_title"] = fresh["lecture_title"]
            e["schedule_row"] = fresh["schedule_row"]

            aid = old.get("assignment_id")
            if aid:
                # look the practice partner up from YOUR chosen assignment,
                # not from whatever this run would have proposed
                aname = ALL.get(aid, old.get("assignment_name") or "")
                e["assignment_name"] = ALL.get(aid, old.get("assignment_name"))
                new_pid = practice_for(aname)
                old_pid = old.get("practice_id")
                if new_pid != old_pid:
                    e["practice_id"] = new_pid
                    e["practice_name"] = P.get(new_pid) if new_pid else None
                    if new_pid and not old_pid:
                        merge_report.append((lid, "PRACTICE ADDED", P[new_pid]))
                    elif old_pid and not new_pid:
                        merge_report.append((lid, "PRACTICE LOST",
                                             f"{old.get('practice_name')} no longer matches"))
                    else:
                        merge_report.append((lid, "PRACTICE CHANGED",
                                             f"{old.get('practice_name')} -> {P.get(new_pid)}"))
                else:
                    e["practice_name"] = P.get(old_pid) if old_pid else None
            merged[lid] = e
        for lid in prev:
            if lid not in merged:
                merged[lid] = prev[lid]
                merge_report.append((lid, "KEPT", "no longer produced by the matcher; left as-is"))
        out = merged
    elif args.merge:
        print(f"# --merge given but {OUT_MAP} does not exist — writing a fresh map\n",
              file=sys.stderr)

    if os.path.exists(OUT_MAP):
        bak = OUT_MAP + ".bak"
        with open(OUT_MAP, encoding="utf-8") as f_in, open(bak, "w") as f_out:
            f_out.write(f_in.read())
        print(f"# previous map backed up to {bak}", file=sys.stderr)

    json.dump(out, open(OUT_MAP, "w"), indent=1)

    if merge_report:
        print("=" * 118)
        print("MERGE — your assignment choices, skips and confidence edits were kept; "
              "practice links refreshed")
        print("=" * 118)
        for lid, kind, detail in merge_report:
            print(f"  {lid:<5} {kind:<17} {detail}")
        print()

    # ---- review table ----
    print("=" * 118)
    print(f"PROPOSED LECTURE -> DAILY HOMEWORK MAP   (review and edit {OUT_MAP})")
    print("=" * 118)
    print(f"{'LEC':<5} {'CONF':<7} {'LECTURE':<44} {'DAILY HOMEWORK':<38} PRACTICE")
    print("-" * 118)
    for lid, e in out.items():
        prac = "yes" if e["practice_id"] else "-"
        print(f"{lid:<5} {e['confidence']:<7} {e['lecture_title'][:43]:<44} "
              f"{(e['assignment_name'] or '—')[:37]:<38} {prac}")

    high = [l for l, e in out.items() if e["confidence"] == "high"]
    review = [l for l, e in out.items() if e["confidence"] == "review"]
    none = [l for l, e in out.items() if e["confidence"] == "none"]
    withp = [l for l, e in out.items() if e["practice_id"]]

    print("\n" + "-" * 118)
    print(f"high confidence: {len(high)}    needs review: {len(review)}    no match: {len(none)}")
    print(f"lectures that will get a practice link: {len(withp)} of {len(out)}")
    print("-" * 118)

    if none:
        print("\nNO MATCH — these pages get nothing unless you fill in an assignment_id by hand:")
        for lid in none:
            print(f"    {lid}  {out[lid]['lecture_title'][:50]:52} {out[lid]['notes']}")
    if review:
        print("\nNEEDS REVIEW:")
        for lid in review:
            print(f"    {lid}  -> {out[lid]['assignment_name'][:45]:47} {out[lid]['notes']}")

    linked = set(claims)
    orphans = [(i, n) for i, n in daily if i not in linked]
    if orphans:
        print(f"\nDAILY HOMEWORK NOT LINKED FROM ANY LECTURE ({len(orphans)}) — "
              f"either a lecture is missing, or one lecture needs two links:")
        for i, n in orphans:
            print(f"    {i:>7}  {n}")

    unused_p = [(i, n) for i, n in practice
                if i not in {e["practice_id"] for e in out.values() if e["practice_id"]}]
    if unused_p:
        print(f"\nPRACTICE ASSIGNMENTS NOT LINKED ({len(unused_p)}) "
              f"(spring ones are expected here):")
        for i, n in unused_p:
            print(f"    {i:>7}  {n}")

    print(f"\nWrote {OUT_MAP}.  Edit it, then run without --propose to preview the pages.")


# ---- HTML block -------------------------------------------------------------
def build_block(course, entry, points_label):
    aid, aname = entry["assignment_id"], entry["assignment_name"]
    pid, pname = entry.get("practice_id"), entry.get("practice_name")
    li = [f'<li><a href="/courses/{course}/assignments/{aid}">'
          f'{html.escape(aname)}</a>{points_label}</li>']
    if pid:
        li.append(f'<li><a href="/courses/{course}/assignments/{pid}">'
                  f'{html.escape(pname)}</a> — extra practice, not graded</li>')
    return (
        f'{START}\n'
        f'<div class="daily-hw-links" style="border:1px solid #c7cdd1;border-left:4px solid #0374b5;'
        f'border-radius:4px;padding:0.75em 1em;margin:1.5em 0;background:#f5f8fa;">\n'
        f'<h3 style="margin-top:0;">Homework for this lecture</h3>\n'
        f'<ul style="margin-bottom:0;">\n' + "\n".join(li) + f'\n</ul>\n</div>\n{END}'
    )


# Canvas STRIPS HTML COMMENTS when it sanitises a page body, so the
# <!-- daily-hw:start --> markers this script used to rely on never survive a
# save. Anchor on the div's class instead — that does survive (Canvas rewrites
# the style shorthand and turns hrefs absolute, but leaves class alone).
BLOCK_RE = re.compile(r'<div[^>]*\bdaily-hw-links\b[^>]*>.*?</div>', re.S | re.I)
COMMENT_RE = re.compile(re.escape(START) + r'.*?' + re.escape(END), re.S)
ASSIGN_ID_RE = re.compile(r'/assignments/(\d+)')


def find_blocks(body):
    """Every homework block currently on the page."""
    return BLOCK_RE.findall(body or "")


def block_ids(fragment):
    """Assignment ids referenced inside one block."""
    return {int(x) for x in ASSIGN_ID_RE.findall(fragment or "")}


def page_is_current(body, entry):
    """True when the page already has exactly one block linking exactly the
    right assignments. Byte comparison is useless here — Canvas normalises the
    HTML on save (style longhand, absolute hrefs, data-api-* attributes), so a
    freshly built block never equals what comes back."""
    blocks = find_blocks(body)
    if len(blocks) != 1:
        return False
    want = {entry["assignment_id"]}
    if entry.get("practice_id"):
        want.add(entry["practice_id"])
    return block_ids(blocks[0]) == want


def strip_blocks(body):
    """Remove every homework block, old comment-wrapped ones included."""
    body = COMMENT_RE.sub("", body or "")
    body = BLOCK_RE.sub("", body)
    return re.sub(r'\n{3,}', '\n\n', body).strip()


def splice(body, block, position):
    """Remove ALL existing homework blocks, then add exactly one.

    Removing all of them is what repairs pages that got a duplicate block from
    the comment-marker bug."""
    cleaned = strip_blocks(body)
    return (block + "\n" + cleaned) if position == "top" else (cleaned + "\n" + block)


# ---- phase 2: apply ---------------------------------------------------------
def apply_map(args):
    if not os.path.exists(OUT_MAP):
        sys.exit(f"ERROR: {OUT_MAP} not found — run with --propose first.")
    mapping = json.load(open(OUT_MAP, encoding="utf-8"))

    todo, skipped = [], []
    for lid, e in mapping.items():
        if e.get("skip"):
            skipped.append((lid, "skip: true in the map file"))
        elif not e.get("assignment_id"):
            skipped.append((lid, "no assignment_id set"))
        elif not e.get("page_url"):
            skipped.append((lid, "no page_url"))
        else:
            todo.append((lid, e))

    label = f" — {args.points:g} pts" if args.points else ""
    mode = "APPLY (writing to Canvas)" if args.apply else "DRY RUN (nothing will be written)"
    print(f"# {mode}: {len(todo)} page(s) to update, {len(skipped)} skipped\n", file=sys.stderr)

    still_review = [l for l, e in mapping.items()
                    if e.get("confidence") == "review" and not e.get("skip")
                    and e.get("assignment_id")]
    if still_review and args.apply and not args.force:
        sys.exit(f"REFUSING TO APPLY: {len(still_review)} entries are still marked "
                 f'confidence "review" ({", ".join(still_review)}).\n'
                 f'Fix them in {OUT_MAP} and set "confidence": "high" (or "skip": true), '
                 f"or re-run with --force to write them as-is.")

    changed, unchanged, failures = 0, 0, []
    for lid, e in todo:
        block = build_block(args.course, e, label)
        try:
            page = api(f"courses/{args.course}/pages/{e['page_url']}")[0]
        except urllib.error.HTTPError as ex:
            failures.append((lid, e["page_url"], ex.code, ex.read().decode("utf-8", "replace")[:160]))
            print(f"  FAIL {lid}: HTTP {ex.code} reading page {e['page_url']}")
            continue

        old = page.get("body") or ""
        n_existing = len(find_blocks(old))

        if page_is_current(old, e):
            unchanged += 1
            print(f"  --   {lid}  already correct, no change")
            continue

        new = splice(old, block, args.position)
        if n_existing == 0:
            action = "insert new block"
        elif n_existing == 1:
            action = "replace the existing block"
        else:
            action = f"REMOVE {n_existing} duplicate blocks, insert 1"

        print(f"  {'WRITE' if args.apply else 'would'} {lid}  ({action})  "
              f"{e['assignment_name'][:40]}"
              f"{'  + practice' if e.get('practice_id') else ''}")
        if not args.apply:
            continue
        try:
            api(f"courses/{args.course}/pages/{e['page_url']}", "PUT",
                {"wiki_page[body]": new})
            changed += 1
            time.sleep(0.1)
        except urllib.error.HTTPError as ex:
            failures.append((lid, e["page_url"], ex.code, ex.read().decode("utf-8", "replace")[:160]))
            print(f"  FAIL {lid}: HTTP {ex.code} writing page")

    if skipped:
        print(f"\nSKIPPED ({len(skipped)}):")
        for lid, why in skipped:
            print(f"    {lid}  {why}")

    print(f"\nPages written: {changed}   already correct: {unchanged}   failed: {len(failures)}")
    if failures:
        for lid, url, code, detail in failures:
            print(f"    {lid}  {url}  HTTP {code}  {detail}")

    if not args.apply:
        print("\n--- sample of the block that would be inserted ---")
        if todo:
            print(build_block(args.course, todo[0][1], label))
        print("\n(dry run — add --apply to write)")
        return

    if args.verify:
        print("\nVERIFYING ...")
        bad = []
        for lid, e in todo:
            page = api(f"courses/{args.course}/pages/{e['page_url']}")[0]
            body = page.get("body") or ""
            blocks = find_blocks(body)
            if len(blocks) != 1:
                bad.append(f"    {lid}: expected exactly 1 homework block, found {len(blocks)}")
                continue
            ids = block_ids(blocks[0])
            want = {e["assignment_id"]}
            if e.get("practice_id"):
                want.add(e["practice_id"])
            if ids != want:
                bad.append(f"    {lid}: block links {sorted(ids)}, expected {sorted(want)}")
        if bad:
            print(f"!! {len(bad)} problem(s):")
            for b in bad:
                print(b)
        else:
            print("All updated pages verified.")


def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--course", default=COURSE_ID, help=f"Canvas course id (default {COURSE_ID})")
    ap.add_argument("--propose", action="store_true",
                    help=f"PHASE 1: build {OUT_MAP} from Canvas + the schedule and print a review table")
    ap.add_argument("--merge", action="store_true",
                    help=f"with --propose: keep the assignment_id / skip / confidence values already "
                         f"in {OUT_MAP} and only refresh the practice links. Use this after adding "
                         f"practice assignments in Canvas so your hand edits survive.")
    ap.add_argument("--apply", action="store_true",
                    help="PHASE 2: write the blocks to Canvas (default is a dry run)")
    ap.add_argument("--verify", action="store_true", help="after --apply, re-read each page and check")
    ap.add_argument("--force", action="store_true",
                    help='apply even for entries still marked confidence "review"')
    ap.add_argument("--position", choices=["top", "bottom"], default="bottom",
                    help="where to put the block on a page that doesn't already have one")
    ap.add_argument("--points", type=float, default=3,
                    help="points shown next to the homework link (default 3; use 0 to omit)")
    args = ap.parse_args()

    if args.propose:
        propose(args)
    else:
        apply_map(args)


if __name__ == "__main__":
    main()